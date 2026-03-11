#!/usr/bin/env python3
"""Export audit Excel data to JSON for the dashboard."""
import json
from openpyxl import load_workbook

wb = load_workbook("/Users/j2z0s7a6/Documents/ia/banco_caja_social_bot/docs/Auditoria_Seguridad_BCS_Bot.xlsx")

# Hallazgos
ws = wb["Hallazgos"]
headers = [cell.value for cell in ws[1]]
findings = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        findings.append(dict(zip(headers, row)))

# Plan
ws2 = wb["Plan de Remediación"]
headers2 = [cell.value for cell in ws2[1]]
plan = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    if row[0]:
        plan.append(dict(zip(headers2, row)))

# Summary
ws3 = wb["Resumen por Normativa"]
headers3 = [cell.value for cell in ws3[1]]
summary = []
for row in ws3.iter_rows(min_row=2, values_only=True):
    if row[0]:
        summary.append(dict(zip(headers3, row)))

data = {"findings": findings, "plan": plan, "summary": summary, "generated": "2026-03-11"}

with open("/Users/j2z0s7a6/Documents/ia/banco_caja_social_bot/docs/data/audit_data.json", "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"Exported: {len(findings)} findings, {len(plan)} plan items, {len(summary)} summary rows")
